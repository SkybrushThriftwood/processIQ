"""Tests for processiq.ingestion.excel_loader."""

from io import BytesIO

import pytest
from openpyxl import Workbook

from processiq.exceptions import ExtractionError
from processiq.ingestion.excel_loader import (
    list_sheets,
    load_excel,
    load_excel_from_bytes,
)


def _create_xlsx(rows: list[list], sheet_name: str = "Sheet1") -> bytes:
    """Create a minimal xlsx file in memory from row data."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


VALID_ROWS = [
    ["step_name", "average_time_hours", "resources_needed", "cost_per_instance"],
    ["Review", 1.5, 2, 50],
    ["Approve", 0.5, 1, 25],
]

ALIAS_ROWS = [
    ["task", "hours", "people"],
    ["Step A", 1.0, 1],
    ["Step B", 2.0, 2],
]


class TestLoadExcel:
    def test_load_from_bytes(self):
        xlsx = _create_xlsx(VALID_ROWS)
        result = load_excel(xlsx, process_name="Excel Test")
        assert result.name == "Excel Test"
        assert len(result.steps) == 2
        assert result.steps[0].step_name == "Review"

    def test_file_not_found(self):
        with pytest.raises(ExtractionError):
            load_excel("/nonexistent/path/file.xlsx")

    def test_empty_sheet(self):
        xlsx = _create_xlsx([])
        with pytest.raises(ExtractionError):
            load_excel(xlsx, process_name="Empty")

    def test_with_column_aliases(self):
        xlsx = _create_xlsx(ALIAS_ROWS)
        result = load_excel(xlsx, process_name="Alias Test")
        assert len(result.steps) == 2
        assert result.steps[0].step_name == "Step A"

    def test_header_detection(self):
        """Header not in row 0: title row, then blank, then headers."""
        rows = [
            ["Process Data Report"],  # Title row (row 0)
            [None, None, None],  # Blank row (row 1)
            ["step_name", "average_time_hours", "resources_needed"],  # Header (row 2)
            ["Step A", 1.0, 1],
            ["Step B", 2.0, 2],
        ]
        xlsx = _create_xlsx(rows)
        result = load_excel(xlsx, process_name="Header Test")
        assert len(result.steps) == 2


class TestLoadExcelFromBytes:
    def test_convenience_function(self):
        xlsx = _create_xlsx(VALID_ROWS)
        result = load_excel_from_bytes(xlsx, process_name="Bytes Test")
        assert len(result.steps) == 2


class TestListSheets:
    def test_returns_sheet_names(self):
        wb = Workbook()
        wb.active.title = "Process Data"
        wb.create_sheet("Constraints")
        wb.create_sheet("Notes")
        buf = BytesIO()
        wb.save(buf)
        sheets = list_sheets(buf.getvalue())
        assert sheets == ["Process Data", "Constraints", "Notes"]

    def test_file_not_found(self):
        with pytest.raises(ExtractionError):
            list_sheets("/nonexistent/path/file.xlsx")
